# ============================================================
# prediction.py
# UMKM AI Business Resilience Prediction System
# Module: Prediction for New SME Instances
# Author: AI Engineer
# Version: 1.0.0
# ============================================================

"""
Prediction module for UMKM Business Resilience Prediction System.

Handles:
    - Single SME prediction
    - Batch predictions from DataFrame
    - Probability output per class
    - Confidence computation
    - Natural language prediction summary
    - Export prediction results to Excel
"""

import os
import warnings
import logging
from typing import Any, Dict, List, Optional, Tuple, Union

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

# Class label ordering
CLASS_ORDER: List[str] = ["Low", "Medium", "High"]


# ─────────────────────────────────────────────
# 1. SINGLE PREDICTION
# ─────────────────────────────────────────────

def predict_single(
    model: Any,
    scaler: Any,
    input_scores: Dict[str, float],
    feature_names: List[str],
    class_names: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """
    Predict Business Resilience for a single SME using construct scores.

    Input:
        Six latent construct scores (each 1.0 – 5.0 Likert mean):
            - DigitalCapabilityScore
            - InnovationCapabilityScore
            - EntrepreneurialOrientationScore
            - OrganizationalAgilityScore
            - ResourceAccessScore
            - EnvironmentalDynamismScore

    Output:
        - predicted_class  : e.g. 'High'
        - probabilities    : {'Low': 0.05, 'Medium': 0.15, 'High': 0.80}
        - confidence       : 0.80 (max probability)
        - business_resilience_score : estimated score (1–5)

    Args:
        model: Fitted sklearn-compatible estimator.
        scaler: Fitted feature scaler (StandardScaler or MinMaxScaler).
        input_scores: Dictionary of feature_name → float value.
        feature_names: Ordered list of feature column names for the model.
        class_names: Ordered class labels.

    Returns:
        Dictionary with prediction details.
    """
    if class_names is None:
        class_names = CLASS_ORDER

    # Validate input keys
    missing_keys = [f for f in feature_names if f not in input_scores]
    if missing_keys:
        raise ValueError(f"Missing required input features: {missing_keys}")

    # Build input array in correct feature order
    X_input = np.array([[input_scores[f] for f in feature_names]])

    # Scale using pre-fitted scaler
    if scaler is not None:
        X_scaled = scaler.transform(X_input)
    else:
        X_scaled = X_input

    # Predict
    pred_label = model.predict(X_scaled)[0]

    # Probabilities
    probabilities: Dict[str, float] = {}
    if hasattr(model, "predict_proba"):
        y_prob = model.predict_proba(X_scaled)[0]
        # Map probabilities to class names using model.classes_
        if hasattr(model, "classes_"):
            for cls, prob in zip(model.classes_, y_prob):
                probabilities[str(cls)] = round(float(prob), 4)
        else:
            for cls, prob in zip(class_names, y_prob):
                probabilities[str(cls)] = round(float(prob), 4)
    else:
        # Binary fallback
        for cls in class_names:
            probabilities[cls] = 1.0 if cls == pred_label else 0.0

    confidence = max(probabilities.values()) if probabilities else 1.0

    # Estimate a continuous resilience score from input scores
    # weighted average: OA has highest theoretical influence
    weights = {
        "DigitalCapabilityScore": 0.15,
        "InnovationCapabilityScore": 0.18,
        "EntrepreneurialOrientationScore": 0.12,
        "OrganizationalAgilityScore": 0.25,
        "ResourceAccessScore": 0.20,
        "EnvironmentalDynamismScore": 0.10,  # negatively affects in theory
    }
    br_score = sum(
        input_scores.get(f, 3.0) * w
        for f, w in weights.items()
        if f in feature_names
    )
    # Normalize to 1–5
    br_score = float(np.clip(br_score, 1.0, 5.0))

    result = {
        "predicted_class": str(pred_label),
        "probabilities": probabilities,
        "confidence": round(confidence, 4),
        "business_resilience_score": round(br_score, 3),
        "input_features": input_scores,
    }

    logger.info(
        f"✅ Prediction: {pred_label} | "
        f"Confidence: {confidence:.2%} | "
        f"BR Score: {br_score:.2f}"
    )
    return result


def format_prediction_report(result: Dict[str, Any], sme_name: str = "SME") -> str:
    """
    Format a single prediction result as a readable text report.

    Args:
        result: Output from predict_single().
        sme_name: Name of the SME for display.

    Returns:
        Formatted multi-line string report.
    """
    pred = result["predicted_class"]
    conf = result["confidence"]
    br_score = result["business_resilience_score"]
    probs = result["probabilities"]

    # Emoji indicators
    emoji_map = {"High": "🟢", "Medium": "🟡", "Low": "🔴"}
    advice_map = {
        "High": (
            "Excellent! This UMKM demonstrates robust organizational capabilities "
            "and is well-equipped to navigate market disruptions. "
            "Recommended: Continue digital investment and maintain agility practices."
        ),
        "Medium": (
            "This UMKM shows adequate resilience but has room for improvement. "
            "Focus areas: Strengthening organizational agility and resource access. "
            "Recommended: Targeted capability development programs."
        ),
        "Low": (
            "This UMKM faces significant resilience gaps that require urgent attention. "
            "Priority areas: Digital capability, innovation investment, and external support access. "
            "Recommended: Government assistance programs and mentorship engagement."
        ),
    }

    report = f"""
╔══════════════════════════════════════════════════════════╗
║         BUSINESS RESILIENCE PREDICTION REPORT            ║
╚══════════════════════════════════════════════════════════╝

SME Name   : {sme_name}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🎯 PREDICTION RESULT
   Category : {emoji_map.get(pred, '⚪')} {pred.upper()}
   BR Score  : {br_score:.2f} / 5.00
   Confidence: {conf:.1%}

📊 CLASS PROBABILITIES
"""
    for cls in CLASS_ORDER:
        prob = probs.get(cls, 0.0)
        bar = "█" * int(prob * 25) + "░" * (25 - int(prob * 25))
        report += f"   {cls:8s} [{bar}] {prob*100:5.1f}%\n"

    report += f"""
📋 INPUT FEATURES
"""
    for feat, val in result["input_features"].items():
        short = feat.replace("Score", "").replace("Capability", " Cap").replace("Orientation", " Orient.")
        report += f"   {short:35s}: {val:.2f}/5.00\n"

    report += f"""
💡 STRATEGIC RECOMMENDATION
   {advice_map.get(pred, '')}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""
    return report


# ─────────────────────────────────────────────
# 2. BATCH PREDICTIONS
# ─────────────────────────────────────────────

def predict_batch(
    model: Any,
    scaler: Any,
    X: Union[pd.DataFrame, np.ndarray],
    feature_names: List[str],
    class_names: Optional[List[str]] = None,
    original_df: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    """
    Generate predictions for a batch of SMEs.

    Args:
        model: Fitted estimator.
        scaler: Fitted feature scaler.
        X: Feature matrix (DataFrame or ndarray).
        feature_names: Feature column names.
        class_names: Class labels.
        original_df: Optional original DataFrame to merge results into.

    Returns:
        DataFrame with predictions, probabilities, and confidence scores
        appended.
    """
    if class_names is None:
        class_names = CLASS_ORDER

    # Convert to numpy if needed
    if isinstance(X, pd.DataFrame):
        X_arr = X[feature_names].values
    else:
        X_arr = X

    # Scale
    if scaler is not None:
        X_scaled = scaler.transform(X_arr)
    else:
        X_scaled = X_arr

    # Predictions
    y_pred = model.predict(X_scaled)

    results_df = pd.DataFrame()

    if original_df is not None:
        results_df = original_df.reset_index(drop=True).copy()

    results_df["Predicted_Category"] = y_pred

    # Probabilities
    if hasattr(model, "predict_proba"):
        y_prob = model.predict_proba(X_scaled)
        # Determine model class order
        model_classes = list(model.classes_) if hasattr(model, "classes_") else class_names

        for cls in class_names:
            if cls in model_classes:
                idx = model_classes.index(cls)
                results_df[f"Prob_{cls}"] = y_prob[:, idx].round(4)
            else:
                results_df[f"Prob_{cls}"] = 0.0

        results_df["Confidence"] = y_prob.max(axis=1).round(4)
    else:
        for cls in class_names:
            results_df[f"Prob_{cls}"] = (y_pred == cls).astype(float)
        results_df["Confidence"] = 1.0

    logger.info(
        f"✅ Batch prediction complete for {len(results_df)} SMEs. "
        f"Distribution: {pd.Series(y_pred).value_counts().to_dict()}"
    )
    return results_df


# ─────────────────────────────────────────────
# 3. EXAMPLE PREDICTION SCENARIOS
# ─────────────────────────────────────────────

def get_example_smes() -> List[Dict[str, Any]]:
    """
    Return example SME profiles for demonstration purposes.

    Each profile includes construct scores (1–5 Likert means)
    and a descriptive business profile.

    Returns:
        List of example SME input dictionaries.
    """
    return [
        {
            "name": "Batik Nusantara Digital (High Resilience)",
            "scores": {
                "DigitalCapabilityScore": 4.6,
                "InnovationCapabilityScore": 4.4,
                "EntrepreneurialOrientationScore": 4.5,
                "OrganizationalAgilityScore": 4.7,
                "ResourceAccessScore": 4.2,
                "EnvironmentalDynamismScore": 3.8,
            },
            "description": (
                "Traditional batik enterprise with strong digital transformation, "
                "active e-commerce presence, and lean agile team management."
            ),
        },
        {
            "name": "Warung Pak Budi (Medium Resilience)",
            "scores": {
                "DigitalCapabilityScore": 2.8,
                "InnovationCapabilityScore": 3.0,
                "EntrepreneurialOrientationScore": 3.2,
                "OrganizationalAgilityScore": 3.1,
                "ResourceAccessScore": 2.9,
                "EnvironmentalDynamismScore": 3.5,
            },
            "description": (
                "Traditional food stall with some digital tools adoption "
                "but facing supply chain and capital access challenges."
            ),
        },
        {
            "name": "Konveksi Tiga Putri (Low Resilience)",
            "scores": {
                "DigitalCapabilityScore": 1.8,
                "InnovationCapabilityScore": 2.0,
                "EntrepreneurialOrientationScore": 1.9,
                "OrganizationalAgilityScore": 1.7,
                "ResourceAccessScore": 1.6,
                "EnvironmentalDynamismScore": 4.2,
            },
            "description": (
                "Small garment manufacturer operating manually with minimal digital "
                "adoption, limited financing access, and high market volatility exposure."
            ),
        },
    ]


# ─────────────────────────────────────────────
# 4. EXPORT PREDICTIONS
# ─────────────────────────────────────────────

def export_predictions(
    predictions_df: pd.DataFrame,
    output_path: str = "prediction_results.xlsx",
    sheet_name: str = "Predictions",
) -> None:
    """
    Export batch prediction results to an Excel file with formatting.

    Args:
        predictions_df: DataFrame from predict_batch().
        output_path: Output file path (.xlsx).
        sheet_name: Excel sheet name.
    """
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        predictions_df.to_excel(writer, index=False, sheet_name=sheet_name)

        # Apply conditional formatting
        try:
            from openpyxl.styles import PatternFill, Font
            ws = writer.sheets[sheet_name]

            # Color-code predicted categories
            color_map = {
                "High": "D5F5E3",    # Light green
                "Medium": "FEF9E7",  # Light yellow
                "Low": "FADBD8",     # Light red
            }

            pred_col_idx = None
            for col_idx, cell in enumerate(ws[1], start=1):
                if cell.value == "Predicted_Category":
                    pred_col_idx = col_idx
                    break

            if pred_col_idx:
                for row in ws.iter_rows(min_row=2, min_col=pred_col_idx, max_col=pred_col_idx):
                    for cell in row:
                        cat = str(cell.value)
                        if cat in color_map:
                            cell.fill = PatternFill(
                                start_color=color_map[cat],
                                end_color=color_map[cat],
                                fill_type="solid",
                            )
                            cell.font = Font(bold=True)
        except Exception as e:
            logger.warning(f"⚠️ Excel formatting skipped: {e}")

    logger.info(f"✅ Predictions exported → {output_path}")


# ─────────────────────────────────────────────
# 5. LOAD MODEL & PREDICT
# ─────────────────────────────────────────────

def predict_from_saved_model(
    model_path: str,
    input_scores: Dict[str, float],
    feature_names: Optional[List[str]] = None,
    sme_name: str = "New SME",
) -> Dict[str, Any]:
    """
    Load a saved model from disk and generate a prediction.

    Convenience function for deployment or standalone usage.

    Args:
        model_path: Path to the .pkl model file.
        input_scores: Dictionary of feature_name → float value.
        feature_names: Optional feature column names.
        sme_name: Display name for the SME.

    Returns:
        Prediction result dictionary.
    """
    import joblib

    payload = joblib.load(model_path)
    model = payload["model"]
    scaler = payload.get("scaler")
    label_encoder = payload.get("label_encoder")

    if feature_names is None:
        feature_names = [
            "DigitalCapabilityScore",
            "InnovationCapabilityScore",
            "EntrepreneurialOrientationScore",
            "OrganizationalAgilityScore",
            "ResourceAccessScore",
            "EnvironmentalDynamismScore",
        ]

    # Recover class names from label encoder if available
    class_names = list(label_encoder.classes_) if label_encoder else CLASS_ORDER

    result = predict_single(model, scaler, input_scores, feature_names, class_names)
    report = format_prediction_report(result, sme_name)
    logger.info(f"\n{report}")

    result["report"] = report
    return result
